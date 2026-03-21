#!/usr/bin/env python3
from __future__ import annotations
"""
Clawmson media handler.
Downloads and processes photos, documents, voice, and audio from Telegram.
Returns (local_path, text_content, has_image) for use in conversation.
"""

import os
import json
import datetime
import requests
from pathlib import Path

_BOT_TOKEN  = ""   # Set by dispatcher via init()
_API_BASE   = ""
MEDIA_PATH  = Path(os.environ.get("CLAWMSON_MEDIA_PATH",
                                   Path.home() / ".openclaw" / "media"))


def init(bot_token: str):
    """Call once from dispatcher after env is loaded."""
    global _BOT_TOKEN, _API_BASE
    _BOT_TOKEN = bot_token
    _API_BASE  = f"https://api.telegram.org/bot{bot_token}"


def _ts_name(prefix: str, ext: str) -> str:
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.{ext}"


def _get_file_url(file_id: str):
    try:
        r = requests.get(f"{_API_BASE}/getFile",
                         params={"file_id": file_id}, timeout=10)
        data = r.json()
        if data.get("ok"):
            fp = data["result"]["file_path"]
            return f"https://api.telegram.org/file/bot{_BOT_TOKEN}/{fp}"
    except Exception:
        pass
    return None


def _download(url: str, dest: Path) -> bool:
    try:
        r = requests.get(url, timeout=120, stream=True)
        r.raise_for_status()
        dest.parent.mkdir(parents=True, exist_ok=True)
        with open(dest, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                f.write(chunk)
        return True
    except Exception:
        return False


# ── Text extraction ──────────────────────────────────────────────────────────

def _extract_text(path: Path, mime: str = "") -> str:
    ext = path.suffix.lower()

    # Plain text / code / config
    if ext in (".txt", ".md", ".py", ".js", ".ts", ".sh", ".bash",
               ".yaml", ".yml", ".toml", ".html", ".css", ".ini", ".cfg"):
        try:
            return path.read_text(errors="replace")[:8000]
        except Exception as e:
            return f"Could not read file: {e}"

    # JSON — parse and pretty-print
    if ext == ".json":
        try:
            data = json.loads(path.read_bytes())
            return json.dumps(data, indent=2)[:6000]
        except Exception:
            try:
                return path.read_text(errors="replace")[:6000]
            except Exception:
                return "(unreadable JSON)"

    # CSV
    if ext == ".csv":
        try:
            import csv
            lines = []
            with open(path, newline="", errors="replace") as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    lines.append(", ".join(row))
                    if i >= 100:
                        lines.append(f"... (truncated at 100 rows)")
                        break
            return "\n".join(lines)[:6000]
        except Exception as e:
            return f"CSV error: {e}"

    # PDF
    if ext == ".pdf" or "pdf" in mime:
        try:
            import pdfplumber
            pages = []
            with pdfplumber.open(str(path)) as pdf:
                for page in pdf.pages[:20]:
                    t = page.extract_text()
                    if t:
                        pages.append(t)
            return "\n\n".join(pages)[:8000] or "(PDF had no extractable text)"
        except ImportError:
            return "(pdfplumber not installed — cannot read PDF)"
        except Exception as e:
            return f"PDF error: {e}"

    # Excel
    if ext in (".xlsx", ".xlsm"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            lines = []
            for sheet_name in wb.sheetnames[:3]:
                ws = wb[sheet_name]
                lines.append(f"=== Sheet: {sheet_name} ===")
                for row in list(ws.iter_rows(values_only=True))[:50]:
                    if any(c is not None for c in row):
                        lines.append("\t".join(str(c or "") for c in row))
            return "\n".join(lines)[:6000]
        except ImportError:
            return "(openpyxl not installed — cannot read Excel file)"
        except Exception as e:
            return f"Excel error: {e}"

    return f"(Unsupported file type: {ext} — saved to {path.name})"


# ── Per-type handlers ────────────────────────────────────────────────────────

def _process_photo(msg: dict) -> tuple:
    """Returns (local_path, description, has_image=True)."""
    MEDIA_PATH.mkdir(parents=True, exist_ok=True)
    photos = msg.get("photo", [])
    if not photos:
        return "", "No photo data found.", True

    # Highest resolution available
    photo  = max(photos, key=lambda p: p.get("file_size", 0))
    url    = _get_file_url(photo["file_id"])
    if not url:
        return "", "Could not retrieve photo.", True

    dest = MEDIA_PATH / _ts_name("photo", "jpg")
    if _download(url, dest):
        caption = msg.get("caption", "")
        return str(dest), f"[Photo saved: {dest.name}]{'. Caption: ' + caption if caption else ''}", True
    return "", "Failed to download photo.", True


def _process_document(msg: dict) -> tuple:
    """Returns (local_path, text_content, has_image)."""
    MEDIA_PATH.mkdir(parents=True, exist_ok=True)
    doc       = msg.get("document", {})
    file_id   = doc.get("file_id")
    file_name = doc.get("file_name", "document")
    mime      = doc.get("mime_type", "")

    if not file_id:
        return "", "No document data.", False

    url = _get_file_url(file_id)
    if not url:
        return "", "Could not retrieve document.", False

    ext  = Path(file_name).suffix.lower().lstrip(".")
    dest = MEDIA_PATH / _ts_name("doc", ext or "bin")
    if not _download(url, dest):
        return str(dest), "Failed to download document.", False

    content   = _extract_text(dest, mime)
    is_image  = ext in ("jpg", "jpeg", "png", "gif", "webp", "bmp")
    return str(dest), content, is_image


def _process_voice(msg: dict) -> tuple:
    """Returns (local_path, transcript, has_image=False)."""
    MEDIA_PATH.mkdir(parents=True, exist_ok=True)
    voice   = msg.get("voice") or msg.get("audio") or {}
    file_id = voice.get("file_id")
    if not file_id:
        return "", "No audio data.", False

    url = _get_file_url(file_id)
    if not url:
        return "", "Could not retrieve audio.", False

    ext  = "ogg" if "voice" in msg else "mp3"
    dest = MEDIA_PATH / _ts_name("audio", ext)
    if not _download(url, dest):
        return str(dest), "Failed to download audio.", False

    transcript = _transcribe(dest)
    return str(dest), transcript, False


def _transcribe(path: Path) -> str:
    try:
        import whisper
        model  = whisper.load_model("base")
        result = model.transcribe(str(path))
        text   = result.get("text", "").strip()
        return text if text else "(empty transcript)"
    except ImportError:
        return "(openai-whisper not installed — cannot transcribe audio. Install with: pip install openai-whisper)"
    except Exception as e:
        return f"Transcription error: {e}"


# ── Public API ───────────────────────────────────────────────────────────────

def handle_message_media(msg: dict) -> tuple:
    """
    Detect and process any media in a Telegram message dict.
    Returns (local_path: str, text_content: str, has_image: bool).
    Returns ("", "", False) if no media present.
    """
    if "photo" in msg:
        return _process_photo(msg)
    if "document" in msg:
        return _process_document(msg)
    if "voice" in msg or "audio" in msg:
        return _process_voice(msg)
    return "", "", False
